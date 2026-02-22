#!/usr/bin/env python3
"""Generate the OMTS Excel import template and example file.

This script produces two files:
  - omts-import-template.xlsx: Empty template with data validation and headers
  - omts-import-example.xlsx: Template populated with the SPEC-001 Section 10 example data

Both files use the multi-sheet structure defined by the Excel Import Format
specification, aligned with the expert panel recommendations.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
METADATA_KEY_FONT = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, num_cols):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def add_data_validation(ws, col_letter, values, prompt_title="", prompt_msg=""):
    """Add a dropdown data validation to a column."""
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = prompt_msg
    dv.promptTitle = prompt_title
    dv.showInputMessage = True
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid value"
    dv.error = f"Must be one of: {', '.join(values)}"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}10000")


def set_col_widths(ws, widths):
    """Set column widths from a dict of {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheet definitions ────────────────────────────────────────────────────────

def create_metadata_sheet(wb):
    """Create the Metadata sheet with file-level fields."""
    ws = wb.active
    ws.title = "Metadata"

    fields = [
        ("Field", "Value", "Description"),
        ("snapshot_date", "", "ISO 8601 date (YYYY-MM-DD) when this snapshot was produced. REQUIRED."),
        ("reporting_entity", "", "ID of the organization node whose perspective this file represents (optional)."),
        ("disclosure_scope", "", "Intended audience: internal, partner, or public (optional)."),
        ("default_confidence", "", "Default data quality confidence: verified, reported, inferred, estimated (optional)."),
        ("default_source", "", "Default data quality source description (optional)."),
        ("default_last_verified", "", "Default date data was last verified, ISO 8601 (optional)."),
    ]

    for row_idx, (field, value, desc) in enumerate(fields, start=1):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=3, value=desc)

        if row_idx == 1:
            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
        else:
            ws.cell(row=row_idx, column=1).font = METADATA_KEY_FONT
            ws.cell(row=row_idx, column=3).alignment = WRAP_ALIGNMENT
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Data validation for disclosure_scope
    dv = DataValidation(type="list", formula1='"internal,partner,public"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B4")

    # Data validation for default_confidence
    dv2 = DataValidation(
        type="list", formula1='"verified,reported,inferred,estimated"', allow_blank=True
    )
    ws.add_data_validation(dv2)
    dv2.add("B5")

    set_col_widths(ws, {"A": 22, "B": 30, "C": 70})
    return ws


def create_organizations_sheet(wb):
    """Create the Organizations sheet."""
    ws = wb.create_sheet("Organizations")

    headers = [
        "id",                       # A - graph-local ID (auto-generated if blank)
        "name",                     # B - REQUIRED
        "jurisdiction",             # C - ISO 3166-1 alpha-2
        "status",                   # D - active/dissolved/merged/suspended
        "lei",                      # E - LEI (20-char)
        "duns",                     # F - DUNS (9-digit)
        "nat_reg_value",            # G - national registry number
        "nat_reg_authority",        # H - GLEIF RA code
        "vat_value",                # I - VAT/tax ID
        "vat_country",              # J - ISO 3166-1 alpha-2
        "internal_id",              # K - internal system ID
        "internal_system",          # L - internal system name (authority)
        "risk_tier",                # M - label: risk-tier
        "kraljic_quadrant",         # N - label: kraljic-quadrant
        "approval_status",          # O - label: approval-status
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    # Data validations
    add_data_validation(ws, "D", ["active", "dissolved", "merged", "suspended"],
                        "Status", "Organization lifecycle status")
    add_data_validation(ws, "M", ["critical", "high", "medium", "low"],
                        "Risk Tier", "General risk classification")
    add_data_validation(ws, "N", ["strategic", "leverage", "bottleneck", "non-critical"],
                        "Kraljic Quadrant", "Kraljic portfolio classification")
    add_data_validation(ws, "O", ["approved", "conditional", "pending", "blocked", "phase-out"],
                        "Approval Status", "Supplier approval status")

    set_col_widths(ws, {
        "A": 18, "B": 30, "C": 14, "D": 12, "E": 24, "F": 14,
        "G": 20, "H": 18, "I": 20, "J": 12, "K": 16, "L": 20,
        "M": 12, "N": 16, "O": 16,
    })
    return ws


def create_facilities_sheet(wb):
    """Create the Facilities sheet."""
    ws = wb.create_sheet("Facilities")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "operator_id",              # C - ref to org id
        "address",                  # D
        "latitude",                 # E
        "longitude",                # F
        "gln",                      # G - GLN (13-digit)
        "internal_id",              # H
        "internal_system",          # I
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))
    set_col_widths(ws, {
        "A": 20, "B": 30, "C": 18, "D": 40, "E": 14, "F": 14,
        "G": 18, "H": 16, "I": 20,
    })
    return ws


def create_goods_sheet(wb):
    """Create the Goods sheet."""
    ws = wb.create_sheet("Goods")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "commodity_code",           # C - HS/CN code
        "unit",                     # D - e.g., kg, mt, pcs
        "gtin",                     # E - GTIN
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))
    set_col_widths(ws, {"A": 20, "B": 30, "C": 16, "D": 10, "E": 18})
    return ws


def create_attestations_sheet(wb):
    """Create the Attestations sheet."""
    ws = wb.create_sheet("Attestations")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "attestation_type",         # C - REQUIRED
        "standard",                 # D
        "issuer",                   # E
        "valid_from",               # F - REQUIRED (YYYY-MM-DD)
        "valid_to",                 # G
        "outcome",                  # H
        "status",                   # I
        "reference",                # J
        "risk_severity",            # K
        "risk_likelihood",          # L
        "attested_entity_id",       # M - which node this attests (for attested_by edge)
        "scope",                    # N - attested_by edge scope
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "C",
                        ["certification", "audit", "due_diligence_statement", "self_declaration", "other"],
                        "Attestation Type", "Type of attestation")
    add_data_validation(ws, "H",
                        ["pass", "conditional_pass", "fail", "pending", "not_applicable"],
                        "Outcome", "Attestation outcome")
    add_data_validation(ws, "I",
                        ["active", "suspended", "revoked", "expired", "withdrawn"],
                        "Status", "Attestation lifecycle status")
    add_data_validation(ws, "K",
                        ["critical", "high", "medium", "low"],
                        "Risk Severity", "Risk severity classification")
    add_data_validation(ws, "L",
                        ["very_likely", "likely", "possible", "unlikely"],
                        "Risk Likelihood", "Risk likelihood")

    set_col_widths(ws, {
        "A": 18, "B": 30, "C": 24, "D": 20, "E": 30, "F": 14,
        "G": 14, "H": 16, "I": 12, "J": 20, "K": 14, "L": 14,
        "M": 22, "N": 20,
    })
    return ws


def create_consignments_sheet(wb):
    """Create the Consignments sheet."""
    ws = wb.create_sheet("Consignments")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "lot_id",                   # C
        "quantity",                 # D
        "unit",                     # E
        "production_date",          # F
        "origin_country",           # G - ISO 3166-1 alpha-2
        "installation_id",          # H - ref to facility
        "direct_emissions_co2e",    # I - CBAM
        "indirect_emissions_co2e",  # J - CBAM
        "emission_factor_source",   # K - CBAM
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "K",
                        ["actual", "default_eu", "default_country"],
                        "Emission Factor", "Source of emissions data")

    set_col_widths(ws, {
        "A": 18, "B": 30, "C": 14, "D": 12, "E": 10, "F": 16,
        "G": 16, "H": 18, "I": 22, "J": 24, "K": 24,
    })
    return ws


def create_supply_relationships_sheet(wb):
    """Create the Supply Relationships sheet for supply-chain edges."""
    ws = wb.create_sheet("Supply Relationships")

    headers = [
        "id",                       # A - edge ID (auto-generated if blank)
        "type",                     # B - supplies/subcontracts/tolls/distributes/brokers/sells_to/operates/produces
        "supplier_id",              # C - source node (supplier/operator/facility)
        "buyer_id",                 # D - target node (buyer/facility/good)
        "valid_from",               # E - REQUIRED (YYYY-MM-DD)
        "valid_to",                 # F
        "commodity",                # G - HS code or description
        "tier",                     # H - tier relative to reporting_entity
        "volume",                   # I
        "volume_unit",              # J
        "annual_value",             # K
        "value_currency",           # L - ISO 4217
        "contract_ref",             # M
        "share_of_buyer_demand",    # N - 0-100
        "service_type",             # O - for distributes edges
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "B",
                        ["supplies", "subcontracts", "tolls", "distributes", "brokers",
                         "sells_to", "operates", "produces", "composed_of"],
                        "Edge Type", "Type of supply/operational relationship")
    add_data_validation(ws, "O",
                        ["warehousing", "transport", "fulfillment", "other"],
                        "Service Type", "For distributes edges only")

    set_col_widths(ws, {
        "A": 14, "B": 16, "C": 18, "D": 18, "E": 14, "F": 14,
        "G": 16, "H": 8, "I": 12, "J": 14, "K": 14, "L": 14,
        "M": 16, "N": 22, "O": 16,
    })
    return ws


def create_corporate_structure_sheet(wb):
    """Create the Corporate Structure sheet for hierarchy edges."""
    ws = wb.create_sheet("Corporate Structure")

    headers = [
        "id",                       # A
        "type",                     # B - ownership/legal_parentage/operational_control/beneficial_ownership
        "subsidiary_id",            # C - source (child/subsidiary/person)
        "parent_id",                # D - target (parent/organization)
        "valid_from",               # E - REQUIRED
        "valid_to",                 # F
        "percentage",               # G - for ownership/beneficial_ownership
        "direct",                   # H - TRUE/FALSE
        "control_type",             # I - for operational_control/beneficial_ownership
        "consolidation_basis",      # J - for legal_parentage
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "B",
                        ["ownership", "legal_parentage", "operational_control", "beneficial_ownership"],
                        "Edge Type", "Type of corporate relationship")
    add_data_validation(ws, "H", ["TRUE", "FALSE"],
                        "Direct", "Direct or indirect relationship")
    add_data_validation(ws, "I",
                        ["franchise", "management", "tolling", "licensed_manufacturing", "other",
                         "voting_rights", "capital", "other_means", "senior_management"],
                        "Control Type", "For operational_control or beneficial_ownership")
    add_data_validation(ws, "J",
                        ["ifrs10", "us_gaap_asc810", "other", "unknown"],
                        "Consolidation Basis", "For legal_parentage only")

    set_col_widths(ws, {
        "A": 14, "B": 24, "C": 18, "D": 18, "E": 14, "F": 14,
        "G": 14, "H": 10, "I": 26, "J": 22,
    })
    return ws


def create_persons_sheet(wb):
    """Create the Persons sheet for beneficial owners and key individuals."""
    ws = wb.create_sheet("Persons")

    headers = [
        "id",                       # A
        "name",                     # B - REQUIRED
        "jurisdiction",             # C - ISO 3166-1 alpha-2
        "role",                     # D
        "nationality",              # E - ISO 3166-1 alpha-2
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))
    set_col_widths(ws, {"A": 18, "B": 30, "C": 14, "D": 20, "E": 14})
    return ws


def create_same_as_sheet(wb):
    """Create the Same As sheet for entity deduplication assertions."""
    ws = wb.create_sheet("Same As")

    headers = [
        "entity_a",                 # A - ref to node ID
        "entity_b",                 # B - ref to node ID
        "confidence",               # C - definite/probable/possible
        "basis",                    # D - justification (e.g., name_match, manual_review)
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "C",
                        ["definite", "probable", "possible"],
                        "Confidence", "Confidence level of the equivalence assertion")

    set_col_widths(ws, {"A": 20, "B": 20, "C": 14, "D": 40})
    return ws


def create_identifiers_sheet(wb):
    """Create the Identifiers sheet for advanced multi-identifier scenarios."""
    ws = wb.create_sheet("Identifiers")

    headers = [
        "node_id",                  # A - ref to node in any sheet
        "scheme",                   # B - lei/duns/gln/nat-reg/vat/internal/extension
        "value",                    # C
        "authority",                # D - required for nat-reg, vat, internal
        "sensitivity",              # E - public/restricted/confidential
        "valid_from",               # F
        "valid_to",                 # G
        "verification_status",      # H
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    style_header_row(ws, len(headers))

    add_data_validation(ws, "B",
                        ["lei", "duns", "gln", "nat-reg", "vat", "internal"],
                        "Scheme", "Identifier scheme")
    add_data_validation(ws, "E",
                        ["public", "restricted", "confidential"],
                        "Sensitivity", "Identifier sensitivity level")
    add_data_validation(ws, "H",
                        ["verified", "reported", "inferred", "unverified"],
                        "Verification", "Verification status")

    set_col_widths(ws, {
        "A": 18, "B": 12, "C": 24, "D": 20, "E": 14, "F": 14,
        "G": 14, "H": 18,
    })
    return ws


def create_readme_sheet(wb):
    """Create a README sheet with instructions."""
    ws = wb.create_sheet("README")

    instructions = [
        ("OMTS Excel Import Template", ""),
        ("", ""),
        ("This workbook is designed for import into the OMTS format using:", ""),
        ("    omtsf import-excel <this-file.xlsx> -o output.omts", ""),
        ("", ""),
        ("SHEET OVERVIEW", ""),
        ("Metadata", "File-level settings: snapshot date, reporting entity, disclosure scope."),
        ("Organizations", "Legal entities (companies, NGOs, government bodies)."),
        ("Facilities", "Physical locations (factories, warehouses, farms, mines)."),
        ("Goods", "Products, materials, or commodities."),
        ("Persons", "Beneficial owners, key individuals (sensitivity: confidential by default)."),
        ("Attestations", "Certifications, audits, due diligence statements."),
        ("Consignments", "Batches, lots, shipments (optional, for CBAM/EUDR)."),
        ("Supply Relationships", "Supply, subcontracting, tolling, distribution edges."),
        ("Corporate Structure", "Ownership, legal parentage, operational control edges."),
        ("Same As", "Entity deduplication: link nodes that represent the same real-world entity."),
        ("Identifiers", "Advanced: additional identifiers beyond the common columns."),
        ("", ""),
        ("REQUIRED FIELDS", ""),
        ("Organizations", "name"),
        ("Facilities", "name"),
        ("Goods", "name"),
        ("Attestations", "name, attestation_type, valid_from"),
        ("Supply Relationships", "type, supplier_id, buyer_id, valid_from"),
        ("Corporate Structure", "type, subsidiary_id, parent_id, valid_from"),
        ("", ""),
        ("AUTO-GENERATED FIELDS", ""),
        ("The import command will auto-generate:", ""),
        ("  - file_salt (cryptographic random)", ""),
        ("  - node/edge IDs (if left blank)", ""),
        ("  - boundary_ref nodes (if disclosure_scope set)", ""),
        ("  - sensitivity defaults per SPEC-004", ""),
        ("", ""),
        ("IDENTIFIER COLUMNS", ""),
        ("Common identifiers have dedicated columns on the Organizations sheet:", ""),
        ("  lei          - Legal Entity Identifier (20-char, validated)", ""),
        ("  duns         - DUNS Number (9-digit, validated)", ""),
        ("  nat_reg_*    - National registry number + GLEIF RA authority code", ""),
        ("  vat_*        - VAT/tax ID + ISO 3166-1 alpha-2 country code", ""),
        ("  internal_*   - Internal system ID + system name", ""),
        ("For multiple IDs of the same scheme, use the Identifiers sheet.", ""),
        ("", ""),
        ("EDGE DIRECTION", ""),
        ("Supply Relationships: supplier_id = who supplies, buyer_id = who buys", ""),
        ("Corporate Structure: subsidiary_id = child entity, parent_id = parent entity", ""),
        ("", ""),
        ("ENTITY DEDUPLICATION", ""),
        ("Use the Same As sheet to link nodes that represent the same real-world entity", ""),
        ("but appear as separate rows (e.g., same company under different names/IDs).", ""),
        ("The import command uses these to generate same_as edges for merge operations.", ""),
        ("", ""),
        ("PERSON NODE PRIVACY", ""),
        ("Person nodes default to confidential sensitivity (SPEC-004).", ""),
        ("If disclosure_scope is 'public', the import command will reject the file", ""),
        ("if any person nodes are present.", ""),
        ("", ""),
        ("SPEC VERSION", ""),
        ("This template targets OMTS spec version 0.1.0", ""),
    ]

    title_font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    normal_font = Font(name="Calibri", size=11)
    code_font = Font(name="Consolas", size=10)

    for row_idx, (col_a, col_b) in enumerate(instructions, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws.cell(row=row_idx, column=2, value=col_b)

        if row_idx == 1:
            cell_a.font = title_font
        elif col_a.isupper() and col_a.strip():
            cell_a.font = section_font
        elif col_a.startswith("    ") or col_a.startswith("  -"):
            cell_a.font = code_font
        else:
            cell_a.font = normal_font

        cell_b.font = normal_font

    set_col_widths(ws, {"A": 50, "B": 60})

    # Move README to first position
    wb.move_sheet("README", offset=-wb.sheetnames.index("README"))
    return ws


# ── Example data ─────────────────────────────────────────────────────────────

def populate_example_data(wb):
    """Populate with SPEC-001 Section 10 example data (Acme-Bolt scenario)."""

    # Metadata
    ws = wb["Metadata"]
    ws["B2"] = "2026-02-17"
    ws["B3"] = "org-acme"
    ws["B4"] = "partner"
    ws["B5"] = "reported"
    ws["B6"] = "manual-review"
    ws["B7"] = "2026-02-17"

    # Organizations
    ws = wb["Organizations"]
    # Row 2: Acme Manufacturing GmbH
    ws["A2"] = "org-acme"
    ws["B2"] = "Acme Manufacturing GmbH"
    ws["C2"] = "DE"
    ws["D2"] = "active"
    ws["E2"] = "5493006MHB84DD0ZWV18"
    ws["F2"] = "081466849"
    ws["G2"] = "HRB86891"
    ws["H2"] = "RA000548"
    ws["I2"] = "DE123456789"
    ws["J2"] = "DE"
    ws["K2"] = "V-100234"
    ws["L2"] = "sap-mm-prod"

    # Row 3: Bolt Supplies Ltd
    ws["A3"] = "org-bolt"
    ws["B3"] = "Bolt Supplies Ltd"
    ws["C3"] = "GB"
    ws["D3"] = "active"
    ws["F3"] = "234567890"
    ws["G3"] = "07228507"
    ws["H3"] = "RA000585"
    ws["M3"] = "low"
    ws["N3"] = "strategic"

    # Facilities
    ws = wb["Facilities"]
    ws["A2"] = "fac-bolt-sheffield"
    ws["B2"] = "Bolt Sheffield Plant"
    ws["C2"] = "org-bolt"
    ws["E2"] = 53.3811
    ws["F2"] = -1.4701

    # Goods
    ws = wb["Goods"]
    ws["A2"] = "good-steel-bolts"
    ws["B2"] = "M10 Steel Hex Bolts"
    ws["C2"] = "7318.15"

    # Attestations
    ws = wb["Attestations"]
    ws["A2"] = "att-sa8000"
    ws["B2"] = "SA8000 Certification"
    ws["C2"] = "certification"
    ws["D2"] = "SA8000:2014"
    ws["E2"] = "Social Accountability International"
    ws["F2"] = "2025-06-01"
    ws["G2"] = "2028-05-31"
    ws["H2"] = "pass"
    ws["I2"] = "active"
    ws["M2"] = "fac-bolt-sheffield"
    ws["N2"] = "working conditions"

    # Supply Relationships
    ws = wb["Supply Relationships"]
    # Edge: Bolt supplies Acme
    ws["A2"] = "edge-001"
    ws["B2"] = "supplies"
    ws["C2"] = "org-bolt"
    ws["D2"] = "org-acme"
    ws["E2"] = "2023-01-15"
    ws["G2"] = "7318.15"
    ws["H2"] = 1

    # Edge: Bolt operates Sheffield plant
    ws["A3"] = "edge-002"
    ws["B3"] = "operates"
    ws["C3"] = "org-bolt"
    ws["D3"] = "fac-bolt-sheffield"
    ws["E3"] = "2018-06-01"

    # Edge: Sheffield produces steel bolts
    ws["A4"] = "edge-003"
    ws["B4"] = "produces"
    ws["C4"] = "fac-bolt-sheffield"
    ws["D4"] = "good-steel-bolts"
    ws["E4"] = "2020-03-01"

    # Corporate Structure
    ws = wb["Corporate Structure"]
    # Edge: Acme owns 51% of Bolt
    ws["A2"] = "edge-004"
    ws["B2"] = "ownership"
    ws["C2"] = "org-acme"
    ws["D2"] = "org-bolt"
    ws["E2"] = "2019-04-01"
    ws["G2"] = 51.0

    # Identifiers (additional - GLN for facility, GTIN for good)
    ws = wb["Identifiers"]
    ws["A2"] = "fac-bolt-sheffield"
    ws["B2"] = "gln"
    ws["C2"] = "5060012340001"
    ws["E2"] = "public"

    ws["A3"] = "fac-bolt-sheffield"
    ws["B3"] = "internal"
    ws["C3"] = "SITE-SHF-01"
    ws["D3"] = "bolt-erp"
    ws["E3"] = "restricted"

    ws["A4"] = "good-steel-bolts"
    ws["B4"] = "org.gs1.gtin"
    ws["C4"] = "05060012340018"
    ws["E4"] = "public"


# ── Supplier List (simplified single-sheet template) ────────────────────────

SUPPLIER_LIST_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
METADATA_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="2F5496")
METADATA_VALUE_FONT = Font(name="Calibri", size=10)
METADATA_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def create_supplier_list_workbook():
    """Create the simplified single-sheet OMTS supplier list workbook.

    Layout:
      Row 1-2: Metadata key-value pairs (reporting entity, snapshot date)
      Row 3:   Blank separator
      Row 4:   Column headers
      Row 5+:  Data rows
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier List"

    # ── Metadata area (rows 1-2) ────────────────────────────────────────────

    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = METADATA_FILL
        ws.cell(row=2, column=col).fill = METADATA_FILL

    ws.cell(row=1, column=1, value="Reporting Entity").font = METADATA_LABEL_FONT
    ws.cell(row=1, column=2).font = METADATA_VALUE_FONT
    ws.cell(row=1, column=3, value="Snapshot Date").font = METADATA_LABEL_FONT
    ws.cell(row=1, column=4).font = METADATA_VALUE_FONT

    ws.cell(row=2, column=1, value="Disclosure Scope").font = METADATA_LABEL_FONT
    ws.cell(row=2, column=2).font = METADATA_VALUE_FONT

    # Data validation for disclosure scope
    dv_scope = DataValidation(
        type="list", formula1='"internal,partner,public"', allow_blank=True
    )
    dv_scope.promptTitle = "Disclosure Scope"
    dv_scope.prompt = "Who will see this file? (default: partner)"
    dv_scope.showInputMessage = True
    ws.add_data_validation(dv_scope)
    dv_scope.add("B2")

    # ── Column headers (row 4) ──────────────────────────────────────────────

    headers = [
        "supplier_name",        # A  REQUIRED
        "jurisdiction",         # B  ISO 3166-1 alpha-2
        "tier",                 # C  1, 2, or 3 (default: 1)
        "parent_supplier",      # D  name of tier N-1 supplier (tier 2/3)
        "commodity",            # E  what they supply
        "valid_from",           # F  relationship start (YYYY-MM-DD)
        "annual_value",         # G
        "value_currency",       # H  ISO 4217
        "contract_ref",         # I
        "lei",                  # J
        "duns",                 # K
        "vat",                  # L
        "vat_country",          # M  ISO 3166-1 alpha-2
        "internal_id",          # N
        "risk_tier",            # O  label
        "kraljic_quadrant",     # P  label
        "approval_status",      # Q  label
        "notes",                # R  free text (not imported into graph)
    ]

    HEADER_ROW = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = SUPPLIER_LIST_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[HEADER_ROW].height = 30
    ws.auto_filter.ref = f"A{HEADER_ROW}:R{HEADER_ROW}"

    # ── Data validations ────────────────────────────────────────────────────

    dv_tier = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    dv_tier.promptTitle = "Tier"
    dv_tier.prompt = "Supply-chain tier: 1 = direct, 2 = sub-supplier, 3 = sub-sub-supplier"
    dv_tier.showInputMessage = True
    dv_tier.showErrorMessage = True
    ws.add_data_validation(dv_tier)
    dv_tier.add("C5:C10000")

    dv_risk = DataValidation(
        type="list", formula1='"critical,high,medium,low"', allow_blank=True
    )
    dv_risk.promptTitle = "Risk Tier"
    dv_risk.prompt = "General risk classification"
    dv_risk.showInputMessage = True
    ws.add_data_validation(dv_risk)
    dv_risk.add("O5:O10000")

    dv_kraljic = DataValidation(
        type="list",
        formula1='"strategic,leverage,bottleneck,non-critical"',
        allow_blank=True,
    )
    dv_kraljic.promptTitle = "Kraljic Quadrant"
    dv_kraljic.prompt = "Kraljic portfolio classification"
    dv_kraljic.showInputMessage = True
    ws.add_data_validation(dv_kraljic)
    dv_kraljic.add("P5:P10000")

    dv_approval = DataValidation(
        type="list",
        formula1='"approved,conditional,pending,blocked,phase-out"',
        allow_blank=True,
    )
    dv_approval.promptTitle = "Approval Status"
    dv_approval.prompt = "Supplier approval status"
    dv_approval.showInputMessage = True
    ws.add_data_validation(dv_approval)
    dv_approval.add("Q5:Q10000")

    # ── Column widths ───────────────────────────────────────────────────────

    set_col_widths(ws, {
        "A": 30, "B": 14, "C": 8, "D": 30, "E": 20, "F": 14,
        "G": 14, "H": 14, "I": 16, "J": 24, "K": 14, "L": 20,
        "M": 14, "N": 16, "O": 12, "P": 18, "Q": 16, "R": 30,
    })

    return wb


def populate_supplier_list_example(wb):
    """Populate the supplier list with a realistic procurement scenario.

    Scenario: Acme Manufacturing's direct and tier-2/3 supplier list for
    steel fastener procurement.
    """
    ws = wb["Supplier List"]

    # Metadata
    ws["B1"] = "Acme Manufacturing GmbH"
    ws["D1"] = "2026-02-22"
    ws["B2"] = "partner"

    # Row 5: Tier 1 — direct supplier
    ws["A5"] = "Bolt Supplies Ltd"
    ws["B5"] = "GB"
    ws["C5"] = 1
    ws["E5"] = "7318.15"
    ws["F5"] = "2023-01-15"
    ws["G5"] = 450000
    ws["H5"] = "EUR"
    ws["I5"] = "MSA-2023-001"
    ws["K5"] = "234567890"
    ws["O5"] = "low"
    ws["P5"] = "strategic"
    ws["Q5"] = "approved"

    # Row 6: Tier 1 — direct supplier
    ws["A6"] = "Nordic Fasteners AB"
    ws["B6"] = "SE"
    ws["C6"] = 1
    ws["E6"] = "7318.15"
    ws["F6"] = "2024-06-01"
    ws["G6"] = 120000
    ws["H6"] = "EUR"
    ws["J6"] = "7317ABCDE1234567890"
    ws["O6"] = "medium"
    ws["P6"] = "leverage"
    ws["Q6"] = "conditional"
    ws["R6"] = "Under evaluation; trial order in progress"

    # Row 7: Tier 1 — direct supplier
    ws["A7"] = "Shanghai Steel Components Co"
    ws["B7"] = "CN"
    ws["C7"] = 1
    ws["E7"] = "7228.70"
    ws["F7"] = "2022-03-01"
    ws["G7"] = 800000
    ws["H7"] = "USD"
    ws["I7"] = "FWA-2022-008"
    ws["N7"] = "V-200891"
    ws["O7"] = "high"
    ws["P7"] = "bottleneck"
    ws["Q7"] = "approved"

    # Row 8: Tier 2 — sub-supplier of Bolt Supplies
    ws["A8"] = "Yorkshire Steel Works"
    ws["B8"] = "GB"
    ws["C8"] = 2
    ws["D8"] = "Bolt Supplies Ltd"
    ws["E8"] = "7208.10"
    ws["F8"] = "2021-09-01"
    ws["O8"] = "low"
    ws["Q8"] = "approved"

    # Row 9: Tier 2 — sub-supplier of Shanghai Steel
    ws["A9"] = "Baosteel Trading Co"
    ws["B9"] = "CN"
    ws["C9"] = 2
    ws["D9"] = "Shanghai Steel Components Co"
    ws["E9"] = "7207.11"
    ws["F9"] = "2020-01-15"
    ws["O9"] = "high"
    ws["R9"] = "Primary raw material supplier for Shanghai Steel"

    # Row 10: Tier 3 — sub-supplier of Baosteel
    ws["A10"] = "Inner Mongolia Mining Corp"
    ws["B10"] = "CN"
    ws["C10"] = 3
    ws["D10"] = "Baosteel Trading Co"
    ws["E10"] = "2601.11"
    ws["O10"] = "critical"
    ws["R10"] = "Iron ore source; LKSG high-risk region"


# ── Main ─────────────────────────────────────────────────────────────────────

def create_workbook():
    """Create the OMTS import workbook with all sheets."""
    wb = Workbook()

    create_metadata_sheet(wb)
    create_organizations_sheet(wb)
    create_facilities_sheet(wb)
    create_goods_sheet(wb)
    create_attestations_sheet(wb)
    create_persons_sheet(wb)
    create_consignments_sheet(wb)
    create_supply_relationships_sheet(wb)
    create_corporate_structure_sheet(wb)
    create_same_as_sheet(wb)
    create_identifiers_sheet(wb)
    create_readme_sheet(wb)

    return wb


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Generate empty full template
    wb_template = create_workbook()
    template_path = os.path.join(script_dir, "omts-import-template.xlsx")
    wb_template.save(template_path)
    print(f"Created: {template_path}")

    # Generate full template with example data
    wb_example = create_workbook()
    populate_example_data(wb_example)
    example_path = os.path.join(script_dir, "omts-import-example.xlsx")
    wb_example.save(example_path)
    print(f"Created: {example_path}")

    # Generate empty supplier list template
    wb_sl_template = create_supplier_list_workbook()
    sl_template_path = os.path.join(script_dir, "omts-supplier-list-template.xlsx")
    wb_sl_template.save(sl_template_path)
    print(f"Created: {sl_template_path}")

    # Generate supplier list with example data
    wb_sl_example = create_supplier_list_workbook()
    populate_supplier_list_example(wb_sl_example)
    sl_example_path = os.path.join(script_dir, "omts-supplier-list-example.xlsx")
    wb_sl_example.save(sl_example_path)
    print(f"Created: {sl_example_path}")


if __name__ == "__main__":
    main()
